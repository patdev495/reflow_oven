import openpyxl
from openpyxl.styles import Font

# === CẤU HÌNH TÊN FILE ===
file_1 = r"F:\dev\reflow\fillter\Thiec.xlsx"  # File chính (nơi cần kiểm tra)
file_2 = r"F:\dev\reflow\fillter\PCB尺寸信息.xlsx"  # File để đối chiếu
output_file = "output_highlighted.xlsx"  # File kết quả

# === ĐỌC FILE 1 & FILE 2 ===
wb1 = openpyxl.load_workbook(file_1)
ws1 = wb1.active  # hoặc wb1["TênSheet"] nếu muốn chỉ định cụ thể

wb2 = openpyxl.load_workbook(file_2)
ws2 = wb2.active

# === TÌM CHỈ SỐ CỘT "Model" ===
def find_column_index(sheet, column_name):
    for col in range(1, sheet.max_column + 1):
        if str(sheet.cell(row=1, column=col).value).strip().lower() == column_name.lower():
            return col
    raise ValueError(f"Không tìm thấy cột '{column_name}' trong sheet.")

col_model_1 = find_column_index(ws1, "Model")
col_model_2 = find_column_index(ws2, "Model")

# === LẤY DANH SÁCH MODEL TỪ FILE 2 ===
models_in_file2 = set()
for row in range(2, ws2.max_row + 1):
    value = ws2.cell(row=row, column=col_model_2).value
    if value is not None:
        models_in_file2.add(str(value).strip())

# === KIỂM TRA & BÔI ĐỎ NHỮNG GIÁ TRỊ KHÔNG TỒN TẠI ===
red_font = Font(color="FF0000")  # Mã màu đỏ

for row in range(2, ws1.max_row + 1):
    cell = ws1.cell(row=row, column=col_model_1)
    value = cell.value
    if value is not None and str(value).strip() not in models_in_file2:
        cell.font = red_font  # Bôi đỏ chữ

# === LƯU KẾT QUẢ ===
wb1.save(output_file)
print(f"✅ Đã xử lý xong! File kết quả lưu tại: {output_file}")
