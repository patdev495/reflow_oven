import openpyxl
from openpyxl.styles import Font

# === CẤU HÌNH FILE ===
file_1 = "output_highlighted.xlsx"   # File cần kiểm tra (chứa cột Model1)
file_2 = r"F:\dev\reflow\fillter\amount_thiec.xlsx"   # File chứa danh sách Model để so sánh
output_file = "output_highlighted1.xlsx"  # File kết quả

# === ĐỌC FILE ===
wb1 = openpyxl.load_workbook(file_1)
ws1 = wb1.active  # Sheet đầu tiên

wb2 = openpyxl.load_workbook(file_2)
ws2 = wb2.active

# === HÀM TÌM CHỈ SỐ CỘT THEO TÊN ===
def find_column_index(sheet, column_name):
    for col in range(1, sheet.max_column + 1):
        header = sheet.cell(row=1, column=col).value
        if header and str(header).strip().lower() == column_name.lower():
            return col
    raise ValueError(f"Không tìm thấy cột '{column_name}' trong sheet.")

# === TÌM CỘT CẦN DÙNG ===
col_model1 = find_column_index(ws1, "Model1")  # Cột Model1 trong file 1
col_model2 = find_column_index(ws2, "Model")   # Cột Model trong file 2

# === LẤY DANH SÁCH MODEL TỪ FILE 2 ===
models = set()
for row in range(2, ws2.max_row + 1):
    val = ws2.cell(row=row, column=col_model2).value
    if val is not None:
        models.add(str(val).strip())

# === DUYỆT CỘT Model1 TRONG FILE 1 VÀ KIỂM TRA ===
yellow_font = Font(color="FFD700")  # Màu chữ vàng (Gold)

for row in range(2, ws1.max_row + 1):
    cell = ws1.cell(row=row, column=col_model1)
    value = cell.value
    if value is not None and str(value).strip() not in models:
        cell.font = yellow_font  # Bôi vàng chữ

# === LƯU KẾT QUẢ ===
wb1.save(output_file)
print(f"✅ Đã xử lý xong! Kết quả lưu tại: {output_file}")
